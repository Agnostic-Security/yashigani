"""
Per-format document fixtures for the sandboxed-extractor parser tests
(plan §6 B1, Tom's extractor slice).

Fixtures are BUILT IN-PROCESS (committed as code, not binary blobs) so they are
reproducible, reviewable, and carry an explicit HIDDEN-data part in each format —
the part the extractor must surface to earn its keep (plan §3.1).

Each builder returns ``bytes`` ready to feed the worker on stdin. The OOXML
builders construct the zip's XML parts directly (the same surface the worker
parses), so the fixture and the parser agree on the part layout without pulling
python-docx/python-pptx into the test env. xlsx uses openpyxl (the worker's xlsx
parser) so the hidden-flag round-trip is faithful. pdf is hand-built + pypdf for
the encrypted case.

Sentinel strings embedded so tests can assert a SPECIFIC hidden value surfaced:
  VISIBLE   — appears in visible body
  HIDDENVAL — appears ONLY in a hidden part (comment/note/hidden cell/etc.)
  METAVAL   — appears ONLY in document metadata
"""
from __future__ import annotations

import io
import zipfile

VISIBLE = "visible-body-alice@example.com"
HIDDENVAL = "hidden-secret-123-45-6789"
METAVAL = "metadata-leaker-SecretAuthor"
#: A sensitive value that lives ONLY in a CUSTOM document property
#: (docProps/custom.xml) — the metadata-only leak the obvious core/app sweep
#: misses. Tests assert it is surfaced (detection) AND absent from any re-render.
CUSTOMVAL = "custom-prop-secret-bob@example.com"

_CONTENT_TYPES = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
    b'<Default Extension="xml" ContentType="application/xml"/>'
    b'<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
    b'</Types>'
)

_RELS = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
)

_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
_P = "http://schemas.openxmlformats.org/presentationml/2006/main"
_CP = "http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
_DC = "http://purl.org/dc/elements/1.1/"


_CUSTOM_NS = "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
_VT_NS = "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"


def _core_props() -> bytes:
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<cp:coreProperties xmlns:cp="{_CP}" xmlns:dc="{_DC}">'
        f'<dc:creator>{METAVAL}</dc:creator>'
        f'<dc:title>Quarterly</dc:title>'
        f'</cp:coreProperties>'
    ).encode()


def _custom_props() -> bytes:
    """docProps/custom.xml carrying a sensitive value in a user-defined property
    (the metadata-only-in-CUSTOM leak the core/app sweep misses)."""
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Properties xmlns="{_CUSTOM_NS}" xmlns:vt="{_VT_NS}">'
        f'<property fmtid="{{D5CDD505-2E9C-101B-9397-08002B2CF9AE}}" '
        f'pid="2" name="ClientContact">'
        f'<vt:lpwstr>{CUSTOMVAL}</vt:lpwstr></property>'
        f'</Properties>'
    ).encode()


def make_docx(*, with_hidden: bool = True) -> bytes:
    """docx with a visible body paragraph, a COMMENT + TRACKED-CHANGE (deleted
    text) hidden part, and document metadata."""
    document = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<w:document xmlns:w="{_W}"><w:body>'
        f'<w:p><w:r><w:t>{VISIBLE}</w:t></w:r></w:p>'
    )
    if with_hidden:
        # A tracked-change DELETION carrying the secret (shipped but "removed").
        document += (
            f'<w:p><w:del><w:r><w:delText>{HIDDENVAL}</w:delText></w:r></w:del></w:p>'
        )
    document += '</w:body></w:document>'

    parts = {
        "[Content_Types].xml": _CONTENT_TYPES,
        "_rels/.rels": _RELS,
        "word/document.xml": document.encode(),
        "docProps/core.xml": _core_props(),
        # CUSTOM property carrying a metadata-only secret (custom.xml leak vector).
        "docProps/custom.xml": _custom_props(),
    }
    if with_hidden:
        # A comment carrying the same secret class — a distinct hidden channel.
        parts["word/comments.xml"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<w:comments xmlns:w="{_W}">'
            f'<w:comment w:id="1"><w:p><w:r><w:t>{HIDDENVAL}</w:t></w:r></w:p></w:comment>'
            f'</w:comments>'
        ).encode()
    return _zip(parts)


def make_xlsx(*, with_hidden: bool = True) -> bytes:
    """xlsx (via openpyxl) with a visible cell, a HIDDEN SHEET + hidden column
    carrying the secret, a defined name, a formula cell, and workbook metadata."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visible"
    ws["A1"] = VISIBLE
    # A formula cell — formula TEXT must surface (data_only=False in the parser).
    ws["B1"] = f'=CONCATENATE("formula-",{1})'
    if with_hidden:
        ws["C1"] = HIDDENVAL
        ws.column_dimensions["C"].hidden = True
        h = wb.create_sheet("Hidden")
        h.sheet_state = "hidden"
        h["A1"] = HIDDENVAL + "-sheet"
        wb.defined_names.add(
            openpyxl.workbook.defined_name.DefinedName(
                "SecretRange", attr_text="Visible!$C$1"
            )
        )
    wb.properties.creator = METAVAL
    # CUSTOM workbook property carrying a metadata-only secret (custom.xml leak).
    try:
        from openpyxl.packaging.custom import StringProperty

        wb.custom_doc_props.append(
            StringProperty(name="ClientContact", value=CUSTOMVAL)
        )
    except Exception:
        pass
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_pptx(*, with_hidden: bool = True) -> bytes:
    """pptx with a visible slide and a SPEAKER-NOTE hidden part + metadata."""
    slide = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<p:sld xmlns:p="{_P}" xmlns:a="{_A}"><p:cSld><p:spTree>'
        f'<p:sp><p:txBody><a:p><a:r><a:t>{VISIBLE}</a:t></a:r></a:p></p:txBody></p:sp>'
        f'</p:spTree></p:cSld></p:sld>'
    )
    parts = {
        "[Content_Types].xml": _CONTENT_TYPES,
        "_rels/.rels": _RELS,
        "ppt/slides/slide1.xml": slide.encode(),
        "docProps/core.xml": _core_props(),
        # CUSTOM property carrying a metadata-only secret (custom.xml leak vector).
        "docProps/custom.xml": _custom_props(),
    }
    if with_hidden:
        notes = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<p:notes xmlns:p="{_P}" xmlns:a="{_A}"><p:cSld><p:spTree>'
            f'<p:sp><p:txBody><a:p><a:r><a:t>{HIDDENVAL}</a:t></a:r></a:p></p:txBody></p:sp>'
            f'</p:spTree></p:cSld></p:notes>'
        )
        parts["ppt/notesSlides/notesSlide1.xml"] = notes.encode()
    return _zip(parts)


# --- pdf -------------------------------------------------------------------

def _minimal_text_pdf(body_text: str, *, title: str) -> bytes:
    """A hand-built single-page PDF with a native text layer + DocInfo title."""
    content = f"BT /F1 12 Tf 72 700 Td ({body_text}) Tj ET".encode()
    objs = [
        b"<</Type/Catalog/Pages 2 0 R>>",
        b"<</Type/Pages/Kids[3 0 R]/Count 1>>",
        b"<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R"
        b"/Resources<</Font<</F1 5 0 R>>>>>>",
        b"<</Length " + str(len(content)).encode() + b">>stream\n" + content + b"\nendstream",
        b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>",
        b"<</Title(" + title.encode() + b")>>",
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets = []
    for i, body in enumerate(objs, start=1):
        offsets.append(len(out))
        out += f"{i} 0 obj".encode() + body + b"endobj\n"
    xref_pos = len(out)
    out += b"xref\n0 " + str(len(objs) + 1).encode() + b"\n"
    out += b"0000000000 65535 f \n"
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (
        b"trailer<</Root 1 0 R/Info 6 0 R/Size " + str(len(objs) + 1).encode() + b">>\n"
        b"startxref\n" + str(xref_pos).encode() + b"\n%%EOF"
    )
    return bytes(out)


def make_pdf_text() -> bytes:
    """A PDF whose page has a NATIVE text layer (parseable) + metadata."""
    return _minimal_text_pdf(f"{VISIBLE} {HIDDENVAL}", title=METAVAL)


def make_pdf_image_only() -> bytes:
    """A PDF page with NO text content stream — stands in for an image-only /
    scanned page. The extractor must emit a needs-OCR signal (fail-closed)."""
    objs = [
        b"<</Type/Catalog/Pages 2 0 R>>",
        b"<</Type/Pages/Kids[3 0 R]/Count 1>>",
        b"<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R>>",
        b"<</Length 0>>stream\n\nendstream",  # empty content stream → no text
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets = []
    for i, body in enumerate(objs, start=1):
        offsets.append(len(out))
        out += f"{i} 0 obj".encode() + body + b"endobj\n"
    xref_pos = len(out)
    out += b"xref\n0 " + str(len(objs) + 1).encode() + b"\n"
    out += b"0000000000 65535 f \n"
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (
        b"trailer<</Root 1 0 R/Size " + str(len(objs) + 1).encode() + b">>\n"
        b"startxref\n" + str(xref_pos).encode() + b"\n%%EOF"
    )
    return bytes(out)


def make_pdf_encrypted() -> bytes:
    """An encrypted PDF — the extractor must fail-closed BLOCK, never crack."""
    import pypdf

    reader = pypdf.PdfReader(io.BytesIO(make_pdf_text()))
    writer = pypdf.PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    writer.encrypt("super-secret-password")
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


# --- shared ----------------------------------------------------------------

def _zip(parts: dict[str, bytes]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, payload in parts.items():
            zf.writestr(name, payload)
    return buf.getvalue()


def truncate(data: bytes, *, keep: float = 0.5) -> bytes:
    """Return a truncated/corrupt copy (malformed → fail-closed test)."""
    return data[: max(8, int(len(data) * keep))]


def ole_encrypted_stub(magic_fmt: bytes = b"") -> bytes:
    """Bytes with the OLE Compound File magic — what Office writes for a
    password-protected OOXML doc. The extractor detects this and fail-closes."""
    return b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512
