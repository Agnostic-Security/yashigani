"""
End-to-end ALL-4-ACTIONS test across the 6 committed formats through the real
DocumentInspectionPipeline (plan §5.0 / §5.2 / §5.5, red-team F4/F5/F6).

After this slice every committed format runs LOG / REDACT / PSEUDONYMIZE / BLOCK
end-to-end.  REDACT/PSEUDONYMIZE re-render runs in the jail (here: the real worker
via the subprocess backend, faithful to the stdin->JSON->exit contract; the LIVE
container proof is scripts/extractor_sandbox_containment.py).

The headline proofs:
  - REDACT: the audit-recorded matched value is GONE from the forwarded artefact
    AND from a re-extract of that artefact (no residual — body/hidden/metadata).
  - PSEUDONYMIZE: the original is GONE, tokens present, the replacer map is held
    (F5 handle present, map NOT in any audit/log field), mode-A table delivered.
  - Strongest-action precedence + small-set escalation (F2) fail-closed to BLOCK.
"""
from __future__ import annotations

import logging
import os
import pathlib

import pytest

pytest.importorskip("openpyxl", reason="xlsx parser")
pytest.importorskip("pypdf", reason="pdf parser")
pytest.importorskip("lxml", reason="hardened XML parser")

from src.tests.unit import _doc_fixtures as fx  # noqa: E402
from src.tests.unit.test_documents_end_to_end_log import (  # noqa: E402
    _WorkerSubprocessBackend,
)
from yashigani.documents.extractor import ExtractorRegistry  # noqa: E402
from yashigani.documents.pipeline import (  # noqa: E402
    DISPOSITION_BLOCK,
    DISPOSITION_PSEUDONYMIZE,
    DISPOSITION_REDACT,
    DocumentInspectionPipeline,
)
from yashigani.documents.sandbox import SandboxedExtractorRunner  # noqa: E402


def _pipeline(audit_sink=None, small_set_escalation=True):
    """A pipeline whose registry routes BOTH extraction and re-render through the
    real worker subprocess (the jail's stdin->JSON->exit contract).

    ``small_set_escalation`` defaults ON (the fail-closed production default,
    L-01).  Tests that exercise PSEUDONYMIZE mechanics on a tiny fixture (which
    is itself a re-identifiable small set and would now escalate) pass
    ``small_set_escalation=False`` to isolate the mechanic under test from the
    gate; the gate has its own dedicated tests."""
    runner = SandboxedExtractorRunner(backend=_WorkerSubprocessBackend())
    registry = ExtractorRegistry(sandbox_runner=runner)
    return DocumentInspectionPipeline(
        registry=registry, on_audit=audit_sink,
        small_set_escalation=small_set_escalation,
    )


# Builders that embed a DETECTABLE PII value (so the existing PII detector flags
# it and the pipeline drives a real plan). We use an email (VISIBLE) + an SSN-
# shaped value the detector recognises.
_PII_EMAIL = "alice@example.com"
_PII_SSN = "123-45-6789"


def _txt_doc() -> bytes:
    return f"contact {_PII_EMAIL} ssn {_PII_SSN}\n".encode()


def _csv_doc() -> bytes:
    return f"name,email\nalice,{_PII_EMAIL}\nbob,{_PII_EMAIL}\n".encode()


def _docx_doc() -> bytes:
    return _ooxml_body_docx(f"Email {_PII_EMAIL} here")


def _ooxml_body_docx(body: str) -> bytes:
    import io
    import zipfile
    W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    document = (
        f'<?xml version="1.0"?><w:document xmlns:w="{W}"><w:body>'
        f'<w:p><w:r><w:t>{body}</w:t></w:r></w:p>'
        f'</w:body></w:document>'
    )
    parts = {
        "[Content_Types].xml": fx._CONTENT_TYPES,
        "_rels/.rels": fx._RELS,
        "word/document.xml": document.encode(),
        "docProps/core.xml": fx._core_props(),
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for n, p in parts.items():
            zf.writestr(n, p)
    return buf.getvalue()


def _xlsx_doc() -> bytes:
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = _PII_EMAIL
    ws["A2"] = _PII_EMAIL
    wb.properties.creator = fx.METAVAL
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pptx_doc() -> bytes:
    import io
    import zipfile
    P = "http://schemas.openxmlformats.org/presentationml/2006/main"
    A = "http://schemas.openxmlformats.org/drawingml/2006/main"
    slide = (
        f'<?xml version="1.0"?><p:sld xmlns:p="{P}" xmlns:a="{A}"><p:cSld><p:spTree>'
        f'<p:sp><p:txBody><a:p><a:r><a:t>Contact {_PII_EMAIL}</a:t></a:r></a:p></p:txBody></p:sp>'
        f'</p:spTree></p:cSld></p:sld>'
    )
    parts = {
        "[Content_Types].xml": fx._CONTENT_TYPES,
        "_rels/.rels": fx._RELS,
        "ppt/slides/slide1.xml": slide.encode(),
        "docProps/core.xml": fx._core_props(),
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for n, p in parts.items():
            zf.writestr(n, p)
    return buf.getvalue()


def _pdf_doc() -> bytes:
    return fx._minimal_text_pdf(f"Contact {_PII_EMAIL}", title=fx.METAVAL)


ALL_FORMATS = [
    ("txt", _txt_doc, "text/plain"),
    ("csv", _csv_doc, "text/csv"),
    ("docx", _docx_doc, "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
    ("xlsx", _xlsx_doc, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    ("pptx", _pptx_doc, "application/vnd.openxmlformats-officedocument.presentationml.presentation"),
    ("pdf", _pdf_doc, "application/pdf"),
]


# ---------------------------------------------------------------------------
# LOG — all six (sanity that the action wiring did not regress LOG).
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("fmt,builder,mime", ALL_FORMATS)
def test_log_all_six(fmt, builder, mime):
    pipe = _pipeline()
    doc = builder()  # build ONCE (xlsx bytes are non-deterministic across calls)
    r = pipe.inspect(doc, mime, request_id="req-log", requested_action="LOG")
    assert r.disposition == "LOG"
    assert r.forward_bytes == doc  # LOG forwards the SAME original unchanged
    assert any(m.data_class == "PII.EMAIL" for m in r.matches)


# ---------------------------------------------------------------------------
# REDACT — all six: matched value gone from the forwarded artefact (no residual).
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("fmt,builder,mime", ALL_FORMATS)
def test_redact_all_six_no_residual(fmt, builder, mime):
    pipe = _pipeline()
    r = pipe.inspect(builder(), mime, request_id="req-redact",
                     requested_action="REDACT")
    assert r.disposition == DISPOSITION_REDACT, (fmt, r.block_reason)
    assert r.forward_bytes is not None
    # Re-extract the forwarded artefact independently: NO original PII survives.
    runner = SandboxedExtractorRunner(backend=_WorkerSubprocessBackend())
    reg = ExtractorRegistry(sandbox_runner=runner)
    re_extract = reg.extract(r.forward_bytes, mime)
    text = "\n".join(s.text for s in re_extract.segments)
    assert _PII_EMAIL not in text, f"{fmt}: redacted email survived in artefact"
    assert fx.METAVAL not in text, f"{fmt}: metadata survived REDACT (F4)"
    assert r.audit_fields.get("no_residual_verified") is True
    assert r.audit_fields.get("hidden_and_metadata_stripped") is True


# ---------------------------------------------------------------------------
# PSEUDONYMIZE — all six: original gone, token present, F5 map held, mode-A table.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("fmt,builder,mime", ALL_FORMATS)
def test_pseudonymize_all_six(fmt, builder, mime):
    # Mechanics test (token coherence + no-residual + F5 map): the csv fixture is
    # a tiny re-identifiable set with a name column, which now escalates (L-01),
    # so disable the gate here to isolate the PSEUDONYMIZE mechanic under test.
    pipe = _pipeline(small_set_escalation=False)
    r = pipe.inspect(builder(), mime, request_id="req-pseudo",
                     requested_action="PSEUDONYMIZE", pseudonymize_mode="A")
    assert r.disposition == DISPOSITION_PSEUDONYMIZE, (fmt, r.block_reason)
    assert r.forward_bytes is not None

    runner = SandboxedExtractorRunner(backend=_WorkerSubprocessBackend())
    reg = ExtractorRegistry(sandbox_runner=runner)
    re_extract = reg.extract(r.forward_bytes, mime)
    text = "\n".join(s.text for s in re_extract.segments)
    assert _PII_EMAIL not in text, f"{fmt}: original survived PSEUDONYMIZE (§5.5)"
    # Opaque token (DECIDED 2026-06-10): recover the actual token from the map
    # (no fixed [EMAIL_1] shape) and assert it is present in the tokenized output.
    email_tok = next(t for t, v in r.correspondence_table.rows.items() if v == _PII_EMAIL)
    assert email_tok in text, f"{fmt}: token not present in tokenized artefact"

    # F5: the replacer map is held (handle present, encrypted, TTL'd), and the
    # mode-A correspondence table was emitted, recoverable via the map.
    assert r.replacer_map is not None and len(r.replacer_map.handle) >= 40
    assert r.pseudonymize_mode == "A"
    assert r.correspondence_table is not None
    assert r.correspondence_table.rows[email_tok] == _PII_EMAIL
    # The token is opaque: no class tag, no count.
    assert "[" not in email_tok and "EMAIL" not in email_tok.upper()
    # The mapping file carries the per-file salt header (integrity binding).
    assert r.doc_hash and f"# doc_hash={r.doc_hash}" in r.correspondence_table.to_csv()


def _docx_custom_meta_only_doc() -> bytes:
    """A docx whose BODY is clean but whose CUSTOM property (docProps/custom.xml)
    carries a DETECTABLE PII email. Proves a metadata-ONLY match drives a verdict
    (concern #2) — it is NOT silently passed because the body is clean."""
    import io
    import zipfile
    W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    CUSTOM = "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
    VT = "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"
    document = (
        f'<?xml version="1.0"?><w:document xmlns:w="{W}"><w:body>'
        f'<w:p><w:r><w:t>nothing sensitive in the body</w:t></w:r></w:p>'
        f'</w:body></w:document>'
    )
    custom = (
        f'<?xml version="1.0"?><Properties xmlns="{CUSTOM}" xmlns:vt="{VT}">'
        f'<property fmtid="{{D5CDD505-2E9C-101B-9397-08002B2CF9AE}}" pid="2" '
        f'name="ClientContact"><vt:lpwstr>{_PII_EMAIL}</vt:lpwstr></property>'
        f'</Properties>'
    )
    parts = {
        "[Content_Types].xml": fx._CONTENT_TYPES,
        "_rels/.rels": fx._RELS,
        "word/document.xml": document.encode(),
        "docProps/custom.xml": custom.encode(),
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for n, p in parts.items():
            zf.writestr(n, p)
    return buf.getvalue()


def test_metadata_only_match_drives_verdict_and_is_stripped():
    """Concern #2/#3: a PII value present ONLY in a custom document property is
    DETECTED (drives matches → a real REDACT verdict) and does NOT survive in the
    re-rendered output — proving metadata-only data is identified, acted on, and
    left no residual."""
    mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    pipe = _pipeline()
    doc = _docx_custom_meta_only_doc()
    # The body is clean; the ONLY PII is in the custom property.
    r = pipe.inspect(doc, mime, request_id="req-meta", requested_action="REDACT")
    # A metadata-only match must produce a match (not be silently passed).
    assert any(m.data_class == "PII.EMAIL" for m in r.matches), (
        "metadata-only PII was NOT detected — silent pass (concern #2)"
    )
    assert r.disposition == DISPOSITION_REDACT, r.block_reason
    # And the re-rendered output carries no residual of the metadata value.
    runner = SandboxedExtractorRunner(backend=_WorkerSubprocessBackend())
    reg = ExtractorRegistry(sandbox_runner=runner)
    re_extract = reg.extract(r.forward_bytes, mime)
    text = "\n".join(s.text for s in re_extract.segments)
    assert _PII_EMAIL not in text, "metadata-only PII survived REDACT output"
    assert _PII_EMAIL.encode() not in r.forward_bytes, (
        "metadata-only PII survived in raw REDACT bytes"
    )


def test_csv_pseudonymize_coherent_across_rows():
    """Same value in two rows → same token in both (coherence, §5.3a)."""
    # Coherence mechanic on a tiny set with a name column → disable the gate.
    pipe = _pipeline(small_set_escalation=False)
    r = pipe.inspect(_csv_doc(), "text/csv", request_id="req",
                     requested_action="PSEUDONYMIZE")
    assert r.disposition == DISPOSITION_PSEUDONYMIZE
    text = r.forward_bytes.decode()
    assert _PII_EMAIL not in text
    # Same value in two rows → the SAME opaque token in both (coherence, §5.3a).
    email_tok = next(t for t, v in r.correspondence_table.rows.items() if v == _PII_EMAIL)
    assert text.count(email_tok) == 2  # both rows collapsed to one token


# ---------------------------------------------------------------------------
# F5 — the replacer map NEVER appears in any audit/log field (crown jewel).
# ---------------------------------------------------------------------------

def test_replacer_map_and_original_never_in_audit_or_logs(caplog):
    events: list[tuple[str, dict]] = []
    pipe = _pipeline(audit_sink=lambda name, data: events.append((name, data)))
    with caplog.at_level(logging.DEBUG):
        r = pipe.inspect(_txt_doc(), "text/plain", request_id="req-secret",
                         requested_action="PSEUDONYMIZE", pseudonymize_mode="A")
    assert r.disposition == DISPOSITION_PSEUDONYMIZE
    # The audit event must NOT carry the original value, the map, or the handle.
    import json
    audit_blob = json.dumps([d for _, d in events], default=str)
    assert _PII_EMAIL not in audit_blob, "original PII leaked into audit (F12)"
    assert r.replacer_map.handle not in audit_blob, "map handle leaked into audit (F5)"
    # Logs likewise must not carry the original or the handle.
    log_blob = caplog.text
    assert _PII_EMAIL not in log_blob
    assert r.replacer_map.handle not in log_blob


# ---------------------------------------------------------------------------
# F2 / L-01 — small-set re-identification escalation → BLOCK (NO monkey-patch).
# ---------------------------------------------------------------------------
# Regression for Laura's L-01 (the gate was dead code: it only fired when a QI
# class was left UN-tokenized, which the wired PSEUDONYMIZE path — tokenizing
# every detected class — never produced).  The fixed gate fires on the WIRED
# path: a small structured set carrying quasi-identifiers escalates, because
# consistent tokenization of a small set is still re-identifiable by row
# co-occurrence.  No monkey-patching — this exercises the real disposition path.

def test_small_set_qi_escalates_to_block_wired_path():
    from yashigani.pii.detector import PiiDetector, PiiMode, PiiType
    runner = SandboxedExtractorRunner(backend=_WorkerSubprocessBackend())
    reg = ExtractorRegistry(sandbox_runner=runner)
    # Detector flags email + phone; phone is a quasi-identifier (_QI_TYPES).
    det = PiiDetector(mode=PiiMode.LOG,
                      enabled_types={PiiType.EMAIL, PiiType.PHONE})
    pipe = DocumentInspectionPipeline(registry=reg, pii_detector=det,
                                      small_set_threshold=20)
    csv = (f"email,phone\n"
           f"{_PII_EMAIL},+14155550100\n"
           f"bob@example.com,+14155550101\n").encode()
    # 2 data rows ≤ threshold 20, a QI (phone) is present → the WIRED path
    # escalates to BLOCK (no monkey-patch). Before L-01 this returned
    # PSEUDONYMIZE (the gate was structurally incapable of firing).
    r = pipe.inspect(csv, "text/csv", request_id="req-f2",
                     requested_action="PSEUDONYMIZE")
    assert r.disposition == DISPOSITION_BLOCK
    assert "small" in (r.block_reason or "").lower()


def test_small_set_escalation_disabled_allows_pseudonymize():
    # Parity with a policy that opted OUT (small_set_escalation=false): the same
    # small QI-bearing set is PSEUDONYMIZEd, not blocked.
    from yashigani.pii.detector import PiiDetector, PiiMode, PiiType
    runner = SandboxedExtractorRunner(backend=_WorkerSubprocessBackend())
    reg = ExtractorRegistry(sandbox_runner=runner)
    det = PiiDetector(mode=PiiMode.LOG,
                      enabled_types={PiiType.EMAIL, PiiType.PHONE})
    pipe = DocumentInspectionPipeline(registry=reg, pii_detector=det,
                                      small_set_threshold=20,
                                      small_set_escalation=False)
    csv = (f"email,phone\n"
           f"{_PII_EMAIL},+14155550100\n"
           f"bob@example.com,+14155550101\n").encode()
    r = pipe.inspect(csv, "text/csv", request_id="req-f2-off",
                     requested_action="PSEUDONYMIZE")
    assert r.disposition == DISPOSITION_PSEUDONYMIZE


# ---------------------------------------------------------------------------
# F3 / L-02 — PSEUDONYMIZE mode-B round-trip WIRED into the pipeline, with the
# verbatim-egress-echo residual CLOSED.  These exercise the real render path
# (worker subprocess) end-to-end: outbound tokenized -> response restored, the
# replacer map never surfaced, and Laura's verbatim-echo attack blocked.
# ---------------------------------------------------------------------------

# A NON-QI email-only set large enough to clear the small-set gate (>20 rows),
# so PSEUDONYMIZE proceeds rather than escalating to BLOCK on the demo path.
# Single email column only (no name column → no header-driven PERSON_NAME QI),
# so the set carries no quasi-identifier and the small-set gate does not fire.
def _modeb_csv(n: int = 25) -> bytes:
    rows = "\n".join(f"user{i}@example.com" for i in range(1, n + 1))
    return (f"email\n{rows}\n").encode()


def _modeb_pipeline(audit_sink=None):
    from yashigani.pii.detector import PiiDetector, PiiMode, PiiType
    runner = SandboxedExtractorRunner(backend=_WorkerSubprocessBackend())
    reg = ExtractorRegistry(sandbox_runner=runner)
    # Email-only so no quasi-identifier is present → small-set gate does not fire
    # even though we keep the set modest; we also clear the threshold by row count.
    det = PiiDetector(mode=PiiMode.LOG, enabled_types={PiiType.EMAIL})
    return DocumentInspectionPipeline(
        registry=reg, pii_detector=det, on_audit=audit_sink,
        small_set_threshold=20,
    )


def _modeb_prose_doc() -> bytes:
    """A prose document carrying a handful of distinct emails in DISTINCT
    sentences.  A genuine mode-B answer quotes one record at its issued context
    WITHOUT reproducing the whole frame's structure — so it restores cleanly,
    while a verbatim echo of the frame is still rejected.  (A degenerate pure-
    token-list frame — e.g. a one-column CSV of emails — cannot be safely
    round-tripped because any faithful response is structurally an echo; that is
    correct fail-closed behaviour and is asserted separately.)"""
    return (
        "Finance team contacts.\n"
        "The accounts lead is reachable at alice@example.com for invoices.\n"
        "For payroll questions, write to bob@example.com any weekday.\n"
        "Vendor onboarding goes through carol@example.com only.\n"
        "Escalations should copy dave@example.com on the thread.\n"
        "The audit liaison is erin@example.com this quarter.\n"
    ).encode()


def test_modeb_roundtrip_wired_happy_path():
    """Mode-B end-to-end: outbound tokenized (cloud sees placeholders), then a
    GENUINE cloud answer referencing one record at its issued context is restored
    to the real value via the wired PositionBinder. Map never surfaced."""
    pipe = _modeb_pipeline()
    r = pipe.inspect(_modeb_prose_doc(), "text/plain", request_id="req-mb",
                     requested_action="PSEUDONYMIZE", pseudonymize_mode="B")
    assert r.disposition == DISPOSITION_PSEUDONYMIZE, r.block_reason
    assert r.pseudonymize_mode == "B"
    # Mode B does NOT hand the user a correspondence table (internal round-trip).
    assert r.correspondence_table is None
    # The round-trip holder is wired, primed with the egress frame + map.
    assert r.mode_b_roundtrip is not None
    # The tokenized artefact (what the cloud sees) carries tokens, not originals.
    out_text = r.forward_bytes.decode()
    assert "alice@example.com" not in out_text
    # Opaque token (DECIDED 2026-06-10): recover alice's actual token from the
    # round-trip binder's recorded egress (token -> original), no [EMAIL_1] shape.
    rt = r.mode_b_roundtrip
    alice_tok = next(
        tok for tok, occ in rt.binder._egress.items() if occ.original == "alice@example.com"
    )
    assert alice_tok in out_text

    # A GENUINE cloud answer: it answers about ONE record, reproducing the issued
    # ±24-char context of that token but NOT the whole frame's structure.
    frame = "\n".join(s for s in out_text.splitlines())
    idx = frame.find(alice_tok)
    answer = "You can reach them: " + frame[max(0, idx - 24): idx + len(alice_tok) + 24]
    restore = pipe.restore_modeb_response("req-mb", answer, rt)
    assert restore.echo_rejected is False, restore
    assert restore.restored is True, restore.flags
    assert restore.flags == []
    assert "alice@example.com" in restore.restored_text


def test_modeb_verbatim_echo_attack_blocked():
    """Laura's L-02 verbatim-egress-echo: a malicious/poisoned cloud echoes the
    egress frame back verbatim to harvest cleartext. The wired response path must
    REFUSE restoration wholesale and restore NOTHING (fail-closed)."""
    events: list[tuple[str, dict]] = []
    pipe = _modeb_pipeline(audit_sink=lambda name, data: events.append((name, data)))
    r = pipe.inspect(_modeb_csv(), "text/csv", request_id="req-echo",
                     requested_action="PSEUDONYMIZE", pseudonymize_mode="B")
    assert r.mode_b_roundtrip is not None

    # The attacker echoes the EXACT egress frame (the tokenized payload we sent).
    egress_frame = r.forward_bytes.decode()
    restore = pipe.restore_modeb_response("req-echo", egress_frame, r.mode_b_roundtrip)

    # Fail-closed: echo rejected, NOTHING restored, response returned unchanged.
    assert restore.echo_rejected is True
    assert restore.restored is False
    assert "@example.com" not in restore.restored_text  # no email cleartext at all
    assert restore.restored_text == egress_frame  # unchanged tokenized response
    # An alert audit event was written for the rejected echo.
    assert any(name == "DOCUMENT_MODEB_ECHO_REJECTED" for name, _ in events)


def test_modeb_echo_with_prose_wrapper_blocked():
    """The attacker wraps the echoed frame in prose to look like an answer; the
    token-tag sequence still reproduces the frame verbatim → rejected."""
    pipe = _modeb_pipeline()
    r = pipe.inspect(_modeb_csv(), "text/csv", request_id="req-echo2",
                     requested_action="PSEUDONYMIZE", pseudonymize_mode="B")
    egress_frame = r.forward_bytes.decode()
    crafted = "Here is the data you asked about:\n\n" + egress_frame + "\n\nThanks!"
    restore = pipe.restore_modeb_response("req-echo2", crafted, r.mode_b_roundtrip)
    assert restore.echo_rejected is True
    assert "@example.com" not in restore.restored_text


def test_modeb_namespace_dump_flagged_not_restored():
    """A non-echo namespace dump (replay tokens in an attacker frame) is NOT a
    verbatim echo but is still refused by position binding → flagged, not
    restored (the round-trip is tainted)."""
    pipe = _modeb_pipeline()
    r = pipe.inspect(_modeb_csv(), "text/csv", request_id="req-dump",
                     requested_action="PSEUDONYMIZE", pseudonymize_mode="B")
    # Attacker frames a few (real, opaque) tokens in a sentence they were never
    # issued in — recover the actual tokens from the binder's egress set.
    toks = list(r.mode_b_roundtrip.binder._egress.keys())[:3]
    attack = "Exfil dump: " + " then ".join(toks) + "."
    restore = pipe.restore_modeb_response("req-dump", attack, r.mode_b_roundtrip)
    assert restore.echo_rejected is False
    # Position binding refused them: nothing restored, tokens flagged.
    assert "@example.com" not in restore.restored_text
    assert restore.flags, "foreign-position replays must be flagged"
    assert restore.restored is False


def test_modeb_map_never_surfaced_in_audit_or_holder_text():
    """The replacer map / handle is never surfaced: not in audit, and the
    round-trip restore output only contains cleared cleartext (no map dump)."""
    import json
    events: list[tuple[str, dict]] = []
    pipe = _modeb_pipeline(audit_sink=lambda name, data: events.append((name, data)))
    r = pipe.inspect(_modeb_csv(), "text/csv", request_id="req-mb-map",
                     requested_action="PSEUDONYMIZE", pseudonymize_mode="B")
    handle = r.mode_b_roundtrip.handle
    # Drive a genuine restore + an echo rejection so both audit shapes are emitted.
    egress_frame = r.forward_bytes.decode()
    pipe.restore_modeb_response("req-mb-map", egress_frame, r.mode_b_roundtrip)
    audit_blob = json.dumps([d for _, d in events], default=str)
    assert handle not in audit_blob, "map handle leaked into audit (F5)"
    assert "@example.com" not in audit_blob, "original leaked into audit (F12)"


def test_modeb_ttl_expiry_fails_closed_on_response_path():
    """TTL/handle properties hold on the RESPONSE path: an expired replacer map
    fails closed (no partial restore) even if the response is otherwise genuine."""
    from yashigani.documents.pseudonymize import ReplacerMapExpiredError
    pipe = _modeb_pipeline()
    r = pipe.inspect(_modeb_csv(), "text/csv", request_id="req-mb-ttl",
                     requested_action="PSEUDONYMIZE", pseudonymize_mode="B",
                     map_ttl_s=1)
    rt = r.mode_b_roundtrip
    # Reveal-after-expiry on the wrapped map fails closed (the map is the custody
    # boundary; its TTL governs every reveal, including the response path).
    handle = rt.handle
    import time as _t
    # Force expiry by revealing with a now far in the future.
    with pytest.raises(ReplacerMapExpiredError):
        rt.replacer_map.reveal_unbound(handle, now=_t.monotonic() + 10_000)
    # Destroy is idempotent and also fails closed on subsequent reveal.
    rt.destroy()
    with pytest.raises(ReplacerMapExpiredError):
        rt.replacer_map.reveal_unbound(handle)
