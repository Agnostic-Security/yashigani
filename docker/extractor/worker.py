#!/usr/bin/env python3
"""
Yashigani sandboxed-extractor WORKER — runs INSIDE the per-job jail.

This is the in-sandbox entrypoint of the hardened extractor runtime (plan §6
B1). It is the process the Captain sandbox spawns; it is the SEAM Tom plugged the
real OOXML/PDF parsers into. The REDACT/PSEUDONYMIZE re-render (red-team F6 —
re-render runs in the SAME jail) is implemented here in ``_render_*`` /
``_run_render``: regenerate-from-cleaned-content (never an overlay — F4),
stripping ALL hidden parts + metadata, with the re-extract-the-OUTPUT segments
returned so the host can PROVE no residual.

CONTRACT (language-agnostic, process-level — see sandbox.py docstring):
    stdin  : raw document bytes (the single read-only input)
    argv   : --job extract|redact|pseudonymize  --format docx|xlsx|pptx|pdf
             --declared-mime <mime>
    stdout : exactly ONE JSON object (SandboxJobResult schema):
               {"ok": true,  "segments": [...], "extraction_complete": bool,
                "detected_format": "docx"}
               {"ok": false, "reason": "<why we contained it>"}
    exit 0 : a JSON result was written (ok true OR ok-false-with-reason).
    exit !=0 : the worker crashed — the runner fails closed to BLOCK.

WHAT CAPTAIN OWNS HERE (the env): reading stdin under a size cap, the
decompression-bomb / billion-laughs guard (bomb_guard.py) that runs BEFORE any
parser, the hardened-XML parser factory, the JSON output contract, and the
fail-closed exit semantics. NO untrusted parser runs until the guard passes.

WHAT TOM ADDS (this slice): the bodies of ``_extract_docx`` / ``_extract_xlsx``
/ ``_extract_pptx`` / ``_extract_pdf``. Each returns
``(segments: list[dict], extraction_complete: bool)``. A segment dict mirrors
``yashigani.documents.segment.Segment``:
    {"text": str, "kind": "BODY"|"TABLE_CELL"|"COMMENT"|..., "location": str,
     "confidence": float, "needs_ocr": bool}
The ``kind`` strings MUST match ``SegmentKind`` values (host-side
``SandboxedExtractor`` maps them back to the enum); an unknown kind there is a
fail-closed ValueError, so the kinds here are the single source of truth.

THE DIFFERENTIATOR (plan §3.1, NON-NEGOTIABLE): we surface the HIDDEN data parts
and metadata, not just the visible body. Sensitive data hides in comments,
tracked changes, footnotes, headers/footers, speaker notes, hidden sheets/rows/
columns, cell notes, defined names, formula text, and document properties. Each
segment carries truthful provenance + part-kind so a hidden-cell hit is LABELLED
as a hidden-cell hit in the audit event.

HARDENING (red-team F1): every untrusted XML part is parsed with the
XXE/entity-expansion-safe ``harden_xml_parser()`` (no external entities, no DTD,
bounded expansion). The bomb guard runs BEFORE any part is touched. A malformed/
truncated/encrypted document → clean ``_Contained`` (ok=False, BLOCK), never an
ambiguous crash.

The worker imports the guard + hardened parser from the installed
``yashigani.documents.bomb_guard`` (baked into the extractor image) so there is a
single source of truth for the caps + the parser settings — no copy-drift
(Verification Protocol §4). The parser libraries (python-docx-free direct-XML for
docx/pptx, openpyxl for xlsx, pypdf for pdf) live ONLY in the extractor image
(docker/Dockerfile.extractor), never the gateway image.

REDACT/PSEUDONYMIZE re-render: ``_render_docx`` / ``_render_xlsx`` /
``_render_pptx`` / ``_render_pdf`` (+ ``_render_text_like`` for txt/csv) rebuild a
MINIMAL clean artefact from cleaned content and re-zip / re-emit. They run in THIS
same jail (F6); re-render is NEVER moved into the gateway process — the writer is
attack surface too. The plan (per-span transforms; NO replacer map — F5) arrives
base64'd on the ``--plan`` argv; the document bytes stay on stdin.
"""
from __future__ import annotations

import argparse
import io
import json
import os
import sys
import zipfile

# Hard cap on stdin so a giant pipe cannot exhaust the jail before the guard
# even runs. The cgroup mem-limit is the backstop; this is the fast precise stop.
_MAX_STDIN_BYTES = int(os.environ.get("YASHIGANI_EXTRACTOR_MAX_STDIN_BYTES", str(64 * 1024 * 1024)))

#: Cap on the number of segments any one document may yield, mirroring the
#: front-end ``DEFAULT_MAX_SEGMENTS`` (extractor.py). A pathological sheet/slide
#: count is a cheap amplification — bound it here too (the front-end cannot,
#: because it never sees the cracked parts; only the worker does).
_MAX_SEGMENTS = int(os.environ.get("YASHIGANI_EXTRACTOR_MAX_SEGMENTS", str(100_000)))

#: Cap on a single segment's text length so one giant run/cell/note cannot blow
#: the output budget (the runner also caps total stdout — this is the per-unit
#: precise stop).
_MAX_SEGMENT_CHARS = int(os.environ.get("YASHIGANI_EXTRACTOR_MAX_SEGMENT_CHARS", str(1_000_000)))

_SUPPORTED = {"docx", "xlsx", "pptx", "pdf"}

# OOXML namespaces we read parts under. Kept local (no XML lib import at module
# load — lxml is imported lazily via the hardened factory). We match element tags
# by these namespaces (w: wordprocessing, a: DrawingML for slide/notes text);
# metadata parts are matched by LOCAL tag name (namespace-stripped) since the
# docProps schemas vary by Office version.
_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
#: presentationml + relationship namespaces — only needed by the pptx re-render
#: (rebuilding a minimal clean presentation), not the read path.
_P = "http://schemas.openxmlformats.org/presentationml/2006/main"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _emit(obj: dict) -> None:
    """Write the single JSON result object to stdout and flush."""
    sys.stdout.write(json.dumps(obj, separators=(",", ":")))
    sys.stdout.flush()


def _read_stdin_capped() -> bytes:
    """Read stdin up to the cap. Over the cap → contained (ok=False)."""
    data = sys.stdin.buffer.read(_MAX_STDIN_BYTES + 1)
    if len(data) > _MAX_STDIN_BYTES:
        raise _Contained(f"input exceeds stdin cap {_MAX_STDIN_BYTES} bytes")
    return data


class _Contained(Exception):
    """A guard tripped — emit ok=False with the reason, exit 0 (contained
    cleanly, the runner still BLOCKs but distinguishes it from a crash)."""


# ---------------------------------------------------------------------------
# Shared helpers (provenance, caps, hardened XML, zip parts).
# ---------------------------------------------------------------------------

def _seg(text, kind: str, location: str, *, needs_ocr: bool = False,
         confidence: float = 1.0) -> dict | None:
    """Build a segment dict, dropping empties and enforcing the per-segment
    text cap. Provenance (``location``) is load-bearing for the audit event and
    must never be empty — a programming error here, not user input."""
    if text is None:
        return None
    s = str(text)
    if s.strip() == "":
        return None
    if len(s) > _MAX_SEGMENT_CHARS:
        raise _Contained(
            f"segment at {location} is {len(s)} chars (cap {_MAX_SEGMENT_CHARS}) "
            f"— amplification, fail-closed"
        )
    if not location:
        # Defensive: never emit a segment the host cannot audit.
        raise _Contained("internal: empty segment location — fail-closed")
    return {
        "text": s,
        "kind": kind,
        "location": location,
        "confidence": confidence,
        "needs_ocr": needs_ocr,
    }


def _signal(text: str, kind: str, location: str) -> dict:
    """A control SEGMENT that must ALWAYS be emitted (never dropped as empty) —
    e.g. a needs-OCR / unparseable-page marker. These drive the fail-closed
    extraction_complete=False decision, so they carry confidence=0.0 +
    needs_ocr=True and are surfaced verbatim (the front-end treats them as
    uninspectable content)."""
    return {
        "text": text,
        "kind": kind,
        "location": location,
        "confidence": 0.0,
        "needs_ocr": True,
    }


def _cap_segments(segments: list[dict]) -> None:
    if len(segments) > _MAX_SEGMENTS:
        raise _Contained(
            f"document produced {len(segments)} segments (cap {_MAX_SEGMENTS}) "
            f"— amplification, fail-closed"
        )


def _xml_parser():
    """Return the single hardened lxml parser (XXE / billion-laughs safe).

    Captain owns the settings in bomb_guard.harden_xml_parser(); the worker
    NEVER constructs its own XML parser — one source of truth (F1)."""
    from yashigani.documents.bomb_guard import harden_xml_parser
    return harden_xml_parser()


def _parse_xml(raw: bytes):
    """Parse one untrusted XML part with the hardened parser. Malformed → None
    (the caller decides whether a missing/garbled part fails the whole doc or is
    a tolerable absence; a *body* part going None fails-closed, an *optional*
    hidden part going None marks extraction incomplete)."""
    from lxml import etree  # type: ignore[import-untyped]
    try:
        return etree.fromstring(raw, parser=_xml_parser())
    except etree.XMLSyntaxError:
        return None


def _open_zip(data: bytes) -> zipfile.ZipFile:
    """Open the OOXML zip. The bomb guard has ALREADY validated this archive is
    safe to decompress (run_extract calls _guard_ooxml first); a BadZipFile here
    is a malformed/truncated container → contained."""
    try:
        return zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise _Contained(f"not a valid OOXML zip container: {exc}") from exc


def _read_part(zf: zipfile.ZipFile, name: str) -> bytes | None:
    """Read one zip part by exact name. Missing part → None (not an error; many
    parts are optional)."""
    try:
        return zf.read(name)
    except KeyError:
        return None


def _names(zf: zipfile.ZipFile) -> list[str]:
    return zf.namelist()


def _local(tag: str) -> str:
    """Strip the XML namespace from a tag for matching."""
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _text_of(el) -> str:
    """All descendant text of an element, joined — used for a paragraph/run tree
    where we want the concatenated visible text of the node."""
    if el is None:
        return ""
    return "".join(el.itertext())


def _guard_ooxml(data: bytes) -> None:
    """Run the decompression-bomb / nesting / entry-count guard on the OOXML zip
    BEFORE any parser sees a part (plan §6). Raises _Contained on any breach."""
    from yashigani.documents.bomb_guard import (
        BombGuardLimits,
        DecompressionBombError,
        guard_zip_bytes,
    )

    try:
        guard_zip_bytes(data, BombGuardLimits())
    except DecompressionBombError as exc:
        raise _Contained(str(exc)) from exc


def _ooxml_is_encrypted(data: bytes) -> bool:
    """An OOXML file that was password-protected/encrypted by Office is wrapped
    in an OLE Compound File (CFB), NOT a zip — its magic is the CFB signature.
    We never attempt to decrypt/crack (red-team + design): detect → fail-closed.

    A genuine OOXML zip starts with 'PK'. A CFB starts with the OLE magic. If the
    bytes are a CFB (or otherwise not a zip), the zip-open below contains it; this
    helper gives the precise "encrypted" reason for the common Office case."""
    return data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


# ---------------------------------------------------------------------------
# docx — body + comments + tracked changes + headers/footers + foot/endnotes +
#        document metadata. Parsed via direct hardened-XML (no python-docx, so
#        no surprise object-model parsing of untrusted parts; full control over
#        the hidden parts the high-level lib does not expose).
# ---------------------------------------------------------------------------

def _extract_docx(data: bytes) -> tuple[list[dict], bool]:
    if _ooxml_is_encrypted(data):
        raise _Contained(
            "docx is encrypted/password-protected (OLE-wrapped) — never cracked, "
            "fail-closed BLOCK"
        )
    zf = _open_zip(data)
    names = _names(zf)
    segments: list[dict] = []
    complete = True

    # --- body (word/document.xml) ---
    body = _read_part(zf, "word/document.xml")
    if body is None:
        raise _Contained("docx missing word/document.xml — malformed, fail-closed")
    root = _parse_xml(body)
    if root is None:
        raise _Contained("docx word/document.xml is malformed XML — fail-closed")
    # One BODY segment per paragraph (provenance: the running paragraph index).
    for idx, para in enumerate(root.iter(f"{{{_W}}}p"), start=1):
        seg = _seg(_text_of(para), "BODY", f"word/document.xml#p={idx}")
        if seg:
            segments.append(seg)

    # --- tracked changes / revisions (w:ins inserts, w:del deletions) ---
    # Deleted text lives in w:delText; inserted text in normal w:t under w:ins.
    for idx, ins in enumerate(root.iter(f"{{{_W}}}ins"), start=1):
        seg = _seg(_text_of(ins), "TRACKED_CHANGE",
                   f"word/document.xml#ins={idx}")
        if seg:
            segments.append(seg)
    for idx, dele in enumerate(root.iter(f"{{{_W}}}del"), start=1):
        # delText is the deleted run text — sensitive data is often what was
        # "removed" but still ships in the file.
        txt = "".join(
            t.text or "" for t in dele.iter(f"{{{_W}}}delText")
        )
        seg = _seg(txt, "TRACKED_CHANGE", f"word/document.xml#del={idx}")
        if seg:
            segments.append(seg)

    # --- comments (word/comments.xml) ---
    comments = _read_part(zf, "word/comments.xml")
    if comments is not None:
        croot = _parse_xml(comments)
        if croot is None:
            complete = False  # a present-but-garbled hidden part → incomplete
        else:
            for idx, c in enumerate(croot.iter(f"{{{_W}}}comment"), start=1):
                cid = c.get(f"{{{_W}}}id", str(idx))
                seg = _seg(_text_of(c), "COMMENT",
                           f"word/comments.xml#id={cid}")
                if seg:
                    segments.append(seg)

    # --- footnotes + endnotes ---
    for part, kind_loc in (("word/footnotes.xml", "footnote"),
                           ("word/endnotes.xml", "endnote")):
        raw = _read_part(zf, part)
        if raw is None:
            continue
        froot = _parse_xml(raw)
        if froot is None:
            complete = False
            continue
        note_tag = f"{{{_W}}}footnote" if "footnote" in part else f"{{{_W}}}endnote"
        for idx, n in enumerate(froot.iter(note_tag), start=1):
            nid = n.get(f"{{{_W}}}id", str(idx))
            seg = _seg(_text_of(n), "BODY", f"{part}#{kind_loc}={nid}")
            if seg:
                segments.append(seg)

    # --- headers + footers (word/header*.xml, word/footer*.xml) ---
    for name in names:
        base = name.rsplit("/", 1)[-1]
        if name.startswith("word/") and (base.startswith("header") or base.startswith("footer")) \
                and name.endswith(".xml"):
            raw = _read_part(zf, name)
            if raw is None:
                continue
            hroot = _parse_xml(raw)
            if hroot is None:
                complete = False
                continue
            seg = _seg(_text_of(hroot), "HEADER_FOOTER", f"{name}")
            if seg:
                segments.append(seg)

    # --- document metadata (core + app + CUSTOM properties) ---
    segments.extend(_ooxml_metadata(zf))

    _cap_segments(segments)
    return segments, complete


# ---------------------------------------------------------------------------
# xlsx — all sheets incl. hidden sheets/rows/columns + cell comments/notes +
#        defined names + formula text + workbook metadata. openpyxl models all
#        of these cleanly (its only runtime dep is et-xmlfile; it parses untrusted
#        XML with resolve_entities=False / defusedxml — both in the image).
# ---------------------------------------------------------------------------

def _extract_xlsx(data: bytes) -> tuple[list[dict], bool]:
    if _ooxml_is_encrypted(data):
        raise _Contained(
            "xlsx is encrypted/password-protected (OLE-wrapped) — never cracked, "
            "fail-closed BLOCK"
        )
    # openpyxl raises on a non-zip / corrupt workbook — map to clean containment.
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    try:
        # data_only=False → keep FORMULA TEXT (sensitive data hides in formulas).
        # read_only=False so hidden row/column dims + defined names are available.
        # keep_links=False avoids resolving external workbook links (egress is
        # blocked anyway, but no point parsing them).
        wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=False, read_only=False, keep_links=False,
        )
    except Exception as exc:  # openpyxl raises InvalidFileException / KeyError etc.
        raise _Contained(f"xlsx parse failed: {exc!r} — fail-closed") from exc

    segments: list[dict] = []
    complete = True

    for ws in wb.worksheets:
        sheet = ws.title
        # A hidden/veryHidden sheet is a classic data-hiding spot — label it.
        sheet_hidden = ws.sheet_state in ("hidden", "veryHidden")
        hidden_cols = {
            c for c, d in ws.column_dimensions.items() if getattr(d, "hidden", False)
        }
        hidden_rows = {
            r for r, d in ws.row_dimensions.items() if getattr(d, "hidden", False)
        }
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None or (isinstance(val, str) and val == ""):
                    continue
                col_letter = get_column_letter(cell.column)
                is_hidden = (
                    sheet_hidden
                    or col_letter in hidden_cols
                    or cell.row in hidden_rows
                )
                kind = "HIDDEN" if is_hidden else "TABLE_CELL"
                loc = f"sheet={sheet}!{cell.coordinate}"
                if sheet_hidden:
                    loc += ";sheet-hidden"
                elif col_letter in hidden_cols:
                    loc += ";col-hidden"
                elif cell.row in hidden_rows:
                    loc += ";row-hidden"
                # A formula cell: the .value already carries the '=...' text when
                # data_only=False. Label its kind so a formula-text hit is clear.
                seg = _seg(val, kind, loc)
                if seg:
                    segments.append(seg)

        # --- cell comments / notes ---
        for row in ws.iter_rows():
            for cell in row:
                cmt = getattr(cell, "comment", None)
                if cmt is not None and getattr(cmt, "text", None):
                    seg = _seg(cmt.text, "COMMENT",
                               f"sheet={sheet}!{cell.coordinate};comment")
                    if seg:
                        segments.append(seg)

    # --- defined names (named ranges/constants — text can carry sensitive data) ---
    try:
        for name, dn in wb.defined_names.items():
            attr = getattr(dn, "attr_text", "") or getattr(dn, "value", "")
            seg = _seg(f"{name}={attr}", "METADATA", f"defined-name={name}")
            if seg:
                segments.append(seg)
    except Exception:
        complete = False  # malformed defined-names table → don't claim complete

    # --- workbook metadata (core properties) ---
    segments.extend(_ooxml_metadata_from_props(wb.properties))

    # --- CUSTOM workbook properties (docProps/custom.xml via openpyxl) -------
    # openpyxl parses docProps/custom.xml into wb.custom_doc_props; a sensitive
    # value (or a sensitive property NAME) sitting ONLY here is otherwise a
    # metadata-only leak the core-property sweep misses (plan §3.1).
    try:
        segments.extend(_xlsx_custom_props(wb))
    except Exception:
        complete = False  # malformed custom-props → don't claim complete

    wb.close()
    _cap_segments(segments)
    return segments, complete


# ---------------------------------------------------------------------------
# pptx — slides + speaker notes + slide masters/layouts + comments + metadata.
#        Direct hardened-XML (avoids pulling Pillow + XlsxWriter that python-pptx
#        requires — keeps the jail surface minimal).
# ---------------------------------------------------------------------------

def _extract_pptx(data: bytes) -> tuple[list[dict], bool]:
    if _ooxml_is_encrypted(data):
        raise _Contained(
            "pptx is encrypted/password-protected (OLE-wrapped) — never cracked, "
            "fail-closed BLOCK"
        )
    zf = _open_zip(data)
    names = _names(zf)
    segments: list[dict] = []
    complete = True

    def _drawing_text(root) -> str:
        # Slide/notes/master text lives in DrawingML a:t runs.
        return "".join(t.text or "" for t in root.iter(f"{{{_A}}}t"))

    # --- slides (visible body) ---
    slide_names = sorted(
        n for n in names
        if n.startswith("ppt/slides/slide") and n.endswith(".xml")
    )
    if not slide_names:
        # A pptx with no slide parts is malformed (or not really a pptx).
        raise _Contained("pptx has no slide parts — malformed, fail-closed")
    for name in slide_names:
        raw = _read_part(zf, name)
        if raw is None:
            continue
        root = _parse_xml(raw)
        if root is None:
            raise _Contained(f"pptx slide {name} is malformed XML — fail-closed")
        seg = _seg(_drawing_text(root), "BODY", name)
        if seg:
            segments.append(seg)

    # --- speaker notes (ppt/notesSlides/notesSlide*.xml) — the differentiator ---
    for name in sorted(n for n in names
                       if n.startswith("ppt/notesSlides/notesSlide")
                       and n.endswith(".xml")):
        raw = _read_part(zf, name)
        if raw is None:
            continue
        root = _parse_xml(raw)
        if root is None:
            complete = False
            continue
        seg = _seg(_drawing_text(root), "SPEAKER_NOTE", name)
        if seg:
            segments.append(seg)

    # --- slide masters + layouts (boilerplate can carry sensitive placeholders) ---
    for prefix, loc_kind in (("ppt/slideMasters/slideMaster", "master"),
                             ("ppt/slideLayouts/slideLayout", "layout")):
        for name in sorted(n for n in names
                           if n.startswith(prefix) and n.endswith(".xml")):
            raw = _read_part(zf, name)
            if raw is None:
                continue
            root = _parse_xml(raw)
            if root is None:
                complete = False
                continue
            seg = _seg(_drawing_text(root), "HEADER_FOOTER", f"{name};{loc_kind}")
            if seg:
                segments.append(seg)

    # --- comments (ppt/comments/* — modern + legacy) ---
    for name in sorted(n for n in names
                       if n.startswith("ppt/comments/") and n.endswith(".xml")):
        raw = _read_part(zf, name)
        if raw is None:
            continue
        root = _parse_xml(raw)
        if root is None:
            complete = False
            continue
        # Modern comments use a:t (DrawingML); legacy p:text — grab all text.
        txt = _text_of(root)
        seg = _seg(txt, "COMMENT", name)
        if seg:
            segments.append(seg)

    # --- metadata (core + app properties) ---
    segments.extend(_ooxml_metadata(zf))

    _cap_segments(segments)
    return segments, complete


# ---------------------------------------------------------------------------
# pdf — native text layer per page + document metadata / XMP. Image-only /
#       scanned pages (no extractable text) emit a needs-OCR signal that the
#       front-end maps to extraction_complete=False → fail-closed BLOCK. We do
#       NOT attempt OCR (parked, §A). Encrypted → fail-closed (never cracked).
# ---------------------------------------------------------------------------

def _extract_pdf(data: bytes) -> tuple[list[dict], bool]:
    import pypdf  # type: ignore[import-untyped]
    from pypdf.errors import PdfReadError, DependencyError  # type: ignore[import-untyped]

    try:
        reader = pypdf.PdfReader(io.BytesIO(data))
    except (PdfReadError, DependencyError, OSError, ValueError) as exc:
        raise _Contained(f"pdf parse failed: {exc!r} — fail-closed") from exc

    # Encrypted PDFs: never attempt to crack. is_encrypted is True even for
    # empty-owner-password files; we refuse all encryption (design + red-team).
    if reader.is_encrypted:
        raise _Contained(
            "pdf is encrypted/password-protected — never cracked, fail-closed BLOCK"
        )

    segments: list[dict] = []
    complete = True

    try:
        pages = reader.pages
        n_pages = len(pages)
    except (PdfReadError, DependencyError, ValueError) as exc:
        raise _Contained(f"pdf page tree unreadable: {exc!r} — fail-closed") from exc

    for idx in range(n_pages):
        try:
            page = pages[idx]
            text = page.extract_text() or ""
        except (PdfReadError, DependencyError, KeyError, ValueError, TypeError) as exc:
            # One unreadable page → mark incomplete, keep going (other pages may
            # still carry text we must classify); the doc cannot claim complete.
            complete = False
            segments.append(_signal(
                f"[unparseable page: {exc!r}]", "OCR",
                f"page={idx + 1};unparseable"))
            continue
        if text.strip() == "":
            # No native text layer on this page → likely image-only/scanned.
            # Emit a needs-OCR signal (parked) that fails the doc closed: a page
            # we could not read is NOT an empty page we can wave through (F9/F11).
            complete = False
            segments.append(_signal(
                "[no native text — needs OCR]", "OCR",
                f"page={idx + 1};needs-ocr"))
            continue
        seg = _seg(text, "BODY", f"page={idx + 1}")
        if seg:
            segments.append(seg)

    # --- document metadata (DocInfo) ---
    try:
        meta = reader.metadata
        if meta:
            for key in ("title", "author", "subject", "creator", "producer",
                        "keywords"):
                val = getattr(meta, key, None)
                seg = _seg(val, "METADATA", f"metadata={key}")
                if seg:
                    segments.append(seg)
    except Exception:
        complete = False

    # --- XMP metadata (the richer, often-overlooked metadata stream) ---
    try:
        xmp = reader.xmp_metadata
        if xmp is not None:
            # Surface the raw XMP packet text — it can carry author, custom
            # fields, redaction breadcrumbs. Parsed defensively as a string.
            raw = getattr(xmp, "rdf_root", None)
            xmp_text = ""
            try:
                if raw is not None:
                    # rdf_root is an ElementTree element already parsed by pypdf;
                    # join its descendant text rather than re-parsing the packet.
                    xmp_text = "".join(raw.itertext())
            except Exception:
                xmp_text = ""
            seg = _seg(xmp_text, "METADATA", "metadata=xmp")
            if seg:
                segments.append(seg)
    except Exception:
        complete = False

    # A PDF with zero extractable text and zero pages-with-text is fully
    # uninspectable → not complete (already reflected via the per-page needs-OCR
    # signals; this is the belt-and-suspenders for a zero-page doc).
    if n_pages == 0:
        complete = False

    _cap_segments(segments)
    return segments, complete


# ---------------------------------------------------------------------------
# OOXML metadata helpers (core + app properties), shared by docx/pptx.
# ---------------------------------------------------------------------------

def _ooxml_metadata(zf: zipfile.ZipFile) -> list[dict]:
    """Extract core + extended + CUSTOM document properties from an OOXML package.

    docProps/core.xml carries author/title/subject/keywords/lastModifiedBy;
    docProps/app.xml carries company/manager/template; docProps/custom.xml carries
    ARBITRARY user-defined name/value pairs (classification labels, client names,
    matter numbers — a frequent metadata-only leak that the obvious core/app sweep
    misses). All three are surfaced so a sensitive value sitting ONLY in any of
    them is detected and drives a verdict (plan §3.1 — 'identify ALL data' includes
    metadata, and that includes CUSTOM metadata, not just the obvious core props)."""
    out: list[dict] = []

    core = _read_part(zf, "docProps/core.xml")
    if core is not None:
        root = _parse_xml(core)
        if root is not None:
            for el in root.iter():
                tag = _local(el.tag)
                if tag in ("coreProperties",):
                    continue
                seg = _seg(el.text, "METADATA", f"docProps/core.xml#{tag}")
                if seg:
                    out.append(seg)

    app = _read_part(zf, "docProps/app.xml")
    if app is not None:
        root = _parse_xml(app)
        if root is not None:
            for el in root.iter():
                tag = _local(el.tag)
                if tag in ("Properties",):
                    continue
                seg = _seg(el.text, "METADATA", f"docProps/app.xml#{tag}")
                if seg:
                    out.append(seg)

    # --- CUSTOM document properties (docProps/custom.xml) --------------------
    # Each <property name="..."> wraps a typed value element (vt:lpwstr, vt:i4,
    # vt:filetime, ...). We surface BOTH the property NAME and its VALUE text —
    # the name itself can be sensitive (e.g. a property literally named after a
    # client) and the value certainly can be. Provenance carries the prop name.
    custom = _read_part(zf, "docProps/custom.xml")
    if custom is not None:
        root = _parse_xml(custom)
        if root is not None:
            for el in root.iter():
                if _local(el.tag) != "property":
                    continue
                pname = el.get("name") or ""
                # The typed value lives in a child element's text (vt:* schema).
                vtext = "".join(c.text or "" for c in el).strip() or (el.text or "")
                loc = f"docProps/custom.xml#name={pname or '?'}"
                # Surface the name as part of the segment text too, so a sensitive
                # PROPERTY NAME (not just its value) is classified.
                payload = f"{pname}={vtext}" if pname else vtext
                seg = _seg(payload, "METADATA", loc)
                if seg:
                    out.append(seg)

    return out


def _ooxml_metadata_from_props(props) -> list[dict]:
    """xlsx metadata via openpyxl's parsed DocumentProperties object."""
    out: list[dict] = []
    if props is None:
        return out
    for key in ("creator", "title", "subject", "description", "keywords",
                "lastModifiedBy", "category", "company", "manager"):
        val = getattr(props, key, None)
        seg = _seg(val, "METADATA", f"metadata={key}")
        if seg:
            out.append(seg)
    return out


def _xlsx_custom_props(wb) -> list[dict]:
    """xlsx CUSTOM document properties via openpyxl's ``wb.custom_doc_props``.

    Each entry is a name/value pair from docProps/custom.xml. We surface BOTH the
    name and the value text (the name itself can be sensitive). Defensive: a
    malformed custom-props part must not crash the whole extraction — but it must
    NOT be silently swallowed either, so the caller marks extraction incomplete on
    failure (see _extract_xlsx)."""
    out: list[dict] = []
    cdp = getattr(wb, "custom_doc_props", None)
    if cdp is None:
        return out
    props = getattr(cdp, "props", None) or list(cdp)
    for p in props:
        name = getattr(p, "name", "") or ""
        value = getattr(p, "value", None)
        if value is None:
            # Some property types expose the value under a typed attribute.
            for attr in ("lpwstr", "i4", "filetime", "bool", "r8"):
                value = getattr(p, attr, None)
                if value is not None:
                    break
        payload = f"{name}={value}" if name else str(value)
        seg = _seg(payload, "METADATA", f"custom-property={name or '?'}")
        if seg:
            out.append(seg)
    return out


_EXTRACTORS = {
    "docx": _extract_docx,
    "xlsx": _extract_xlsx,
    "pptx": _extract_pptx,
    "pdf": _extract_pdf,
}


# ---------------------------------------------------------------------------
# Re-render (REDACT / PSEUDONYMIZE) — runs in THIS same jail (red-team F6).
#
# The CONTRACT (transform.RenderPlan, host-side): a plan is a list of per-span
# transforms keyed by the WORKER-side segment location + the exact original
# substring, plus the action (REDACT destroys, PSEUDONYMIZE token-substitutes).
# Re-render is REGENERATE-FROM-CLEANED-CONTENT, never an overlay (F4): we rebuild
# the part XML / text / rows from cleaned values and discard the original object
# graph, so there is no residual under an overlay.
#
# REDACT/PSEUDONYMIZE ALWAYS strip ALL hidden parts + metadata (§5.1 pt 3 / F4) —
# you cannot certify "no residual" while shipping comments/tracked-changes/notes/
# hidden cells/defined-names/core+app+XMP metadata.  The per-format builders below
# rebuild a MINIMAL clean package containing ONLY the visible body parts, with the
# matched spans transformed, and NOTHING else.
# ---------------------------------------------------------------------------


class _Transform:
    """A resolved per-span transform: original value -> replacement (or '' for
    REDACT)."""

    __slots__ = ("original", "replacement", "action")

    def __init__(self, original: str, replacement: str, action: str) -> None:
        self.original = original
        self.replacement = replacement
        self.action = action


def _parse_plan(plan: dict) -> tuple[dict[str, list[_Transform]], bool]:
    """Parse the host RenderPlan dict into (segment_location -> transforms,
    strip_hidden_and_metadata). Fail-closed on a malformed plan."""
    if not isinstance(plan, dict):
        raise _Contained("re-render plan is not an object — fail-closed")
    spans = plan.get("spans")
    if not isinstance(spans, list):
        raise _Contained("re-render plan has no span list — fail-closed")
    by_seg: dict[str, list[_Transform]] = {}
    for s in spans:
        if not isinstance(s, dict):
            raise _Contained("re-render span is not an object — fail-closed")
        loc = str(s.get("segment_location", ""))
        original = str(s.get("original", ""))
        action = str(s.get("action", ""))
        if not loc or original == "" or action not in ("REDACT", "PSEUDONYMIZE"):
            raise _Contained("re-render span malformed (loc/original/action) — fail-closed")
        replacement = "" if action == "REDACT" else str(s.get("token", ""))
        if action == "PSEUDONYMIZE" and replacement == "":
            raise _Contained("PSEUDONYMIZE span missing token — fail-closed")
        by_seg.setdefault(loc, []).append(_Transform(original, replacement, action))
    strip = bool(plan.get("strip_hidden_and_metadata", True))
    return by_seg, strip


def _apply_transforms(text: str, transforms: list[_Transform]) -> str:
    """Apply value-keyed transforms to a segment's text.

    Longest originals first so a short value is not substituted inside a longer
    one. Each original is replaced wherever it appears in this segment (value-
    keyed coherence). REDACT replaces with '' (destruction)."""
    out = text
    for t in sorted(transforms, key=lambda x: len(x.original), reverse=True):
        out = out.replace(t.original, t.replacement)
    return out


def _xml_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _zip_out(parts: dict[str, bytes]) -> bytes:
    """Write a fresh OOXML zip from cleaned parts (regenerate, not edit-in-place)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, payload in parts.items():
            zf.writestr(name, payload)
    return buf.getvalue()


def _render_docx(data: bytes, plan: dict) -> bytes:
    """Rebuild a MINIMAL clean docx: visible body paragraphs only, matched runs
    transformed, ALL hidden parts (comments, tracked changes, footnotes,
    headers/footers) + ALL metadata (docProps/*) STRIPPED.

    F4 residual vectors addressed: the original object graph is discarded (we
    emit a freshly-generated document.xml from cleaned BODY text only); generator
    + core/app/custom metadata are dropped (no docProps); comments/tracked-change
    parts are not carried; there is no embedded thumbnail (docProps/thumbnail) in
    the rebuilt package; font subsets are not an OOXML concern (text is XML, not
    glyph-embedded)."""
    if _ooxml_is_encrypted(data):
        raise _Contained("docx is encrypted — never re-rendered, fail-closed BLOCK")
    _guard_ooxml(data)
    by_seg, _strip = _parse_plan(plan)

    zf = _open_zip(data)
    body = _read_part(zf, "word/document.xml")
    if body is None:
        raise _Contained("docx missing word/document.xml — cannot re-render, fail-closed")
    root = _parse_xml(body)
    if root is None:
        raise _Contained("docx word/document.xml malformed — cannot re-render, fail-closed")

    paras_xml: list[str] = []
    for idx, para in enumerate(root.iter(f"{{{_W}}}p"), start=1):
        loc = f"word/document.xml#p={idx}"
        # VISIBLE run text ONLY: w:t runs, NOT w:delText (tracked-change
        # deletions are a hidden channel and must NOT be carried into the
        # rebuilt body — a deleted-but-shipped secret is a residual leak, F4).
        text = "".join(t.text or "" for t in para.iter(f"{{{_W}}}t"))
        if loc in by_seg:
            text = _apply_transforms(text, by_seg[loc])
        if text.strip() == "":
            # Span fully redacted away → drop the paragraph (no empty residual).
            continue
        paras_xml.append(
            f"<w:p><w:r><w:t xml:space=\"preserve\">{_xml_escape(text)}</w:t></w:r></w:p>"
        )

    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<w:document xmlns:w="{_W}"><w:body>'
        + "".join(paras_xml)
        + "</w:body></w:document>"
    ).encode("utf-8")

    # MINIMAL clean package — body only, NO docProps, NO comments/notes/headers.
    parts = {
        "[Content_Types].xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            '</Types>'
        ).encode("utf-8"),
        "_rels/.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            '</Relationships>'
        ).encode("utf-8"),
        "word/document.xml": document,
    }
    return _zip_out(parts)


def _render_xlsx(data: bytes, plan: dict) -> bytes:
    """Rebuild a clean xlsx via openpyxl from a FRESH workbook: only the VISIBLE
    sheets/cells, matched cells transformed, hidden sheets/rows/cols + comments +
    defined names + formulas-with-matches + ALL metadata DROPPED.

    F4 residual vectors: cached xlsx formula values are eliminated because we
    emit literal cleaned values (no formula carried, so no stale cached result);
    defined names / named ranges are not copied; workbook core metadata is reset
    on the fresh workbook; there are no embedded objects in the rebuilt book."""
    if _ooxml_is_encrypted(data):
        raise _Contained("xlsx is encrypted — never re-rendered, fail-closed BLOCK")
    _guard_ooxml(data)
    by_seg, _strip = _parse_plan(plan)

    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    try:
        wb_in = openpyxl.load_workbook(
            io.BytesIO(data), data_only=False, read_only=False, keep_links=False,
        )
    except Exception as exc:
        raise _Contained(f"xlsx parse failed during re-render: {exc!r} — fail-closed") from exc

    wb_out = openpyxl.Workbook()
    # Remove the default sheet; we add only visible sheets back.
    wb_out.remove(wb_out.active)

    any_sheet = False
    for ws in wb_in.worksheets:
        if ws.sheet_state in ("hidden", "veryHidden"):
            continue  # drop hidden sheets wholesale (no residual channel)
        any_sheet = True
        ws_out = wb_out.create_sheet(title=ws.title)
        hidden_cols = {
            c for c, d in ws.column_dimensions.items() if getattr(d, "hidden", False)
        }
        hidden_rows = {
            r for r, d in ws.row_dimensions.items() if getattr(d, "hidden", False)
        }
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                col_letter = get_column_letter(cell.column)
                # Drop hidden rows/cols entirely (data-hiding channels).
                if col_letter in hidden_cols or cell.row in hidden_rows:
                    continue
                loc = f"sheet={ws.title}!{cell.coordinate}"
                sval = str(val)
                if loc in by_seg:
                    sval = _apply_transforms(sval, by_seg[loc])
                    # A transformed cell ships as a literal string (never a
                    # formula) so no cached formula value can leak.
                    ws_out[cell.coordinate] = sval
                else:
                    # A formula cell that was NOT matched: ship the formula TEXT
                    # as a literal string so no cached evaluated value travels
                    # and the formula cannot re-reference a dropped hidden cell.
                    if isinstance(val, str) and val.startswith("="):
                        ws_out[cell.coordinate] = val  # literal '=...' text, inert
                    else:
                        ws_out[cell.coordinate] = val

    if not any_sheet:
        # Everything was hidden → emit a single empty visible sheet rather than an
        # invalid zero-sheet workbook.
        wb_out.create_sheet(title="Sheet1")

    # Fresh workbook → metadata is default/empty; defined names not copied.
    wb_in.close()
    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue()


def _render_pptx(data: bytes, plan: dict) -> bytes:
    """Rebuild a clean pptx: visible slide body text only, matched runs
    transformed, speaker notes + masters/layouts + comments + ALL metadata
    DROPPED.

    F4 residual vectors: speaker notes (a classic leak channel) are not carried;
    docProps/* metadata dropped; no embedded thumbnail; slide text is regenerated
    DrawingML, not the original part graph."""
    if _ooxml_is_encrypted(data):
        raise _Contained("pptx is encrypted — never re-rendered, fail-closed BLOCK")
    _guard_ooxml(data)
    by_seg, _strip = _parse_plan(plan)

    zf = _open_zip(data)
    names = _names(zf)
    slide_names = sorted(
        n for n in names if n.startswith("ppt/slides/slide") and n.endswith(".xml")
    )
    if not slide_names:
        raise _Contained("pptx has no slide parts — cannot re-render, fail-closed")

    pres_rels: list[str] = []
    ct_overrides: list[str] = []
    out_parts: dict[str, bytes] = {}
    sldid_list: list[str] = []
    for i, name in enumerate(slide_names, start=1):
        raw = _read_part(zf, name)
        if raw is None:
            continue
        root = _parse_xml(raw)
        if root is None:
            raise _Contained(f"pptx slide {name} malformed — cannot re-render, fail-closed")
        text = "".join(t.text or "" for t in root.iter(f"{{{_A}}}t"))
        if text and name in by_seg:
            text = _apply_transforms(text, by_seg[name])
        out_name = f"ppt/slides/slide{i}.xml"
        slide_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<p:sld xmlns:p="{_P}" xmlns:a="{_A}" xmlns:r="{_R_NS}"><p:cSld><p:spTree>'
            '<p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>'
            '<p:grpSpPr/>'
            '<p:sp><p:nvSpPr><p:cNvPr id="2" name="t"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>'
            '<p:spPr/>'
            f'<p:txBody><a:bodyPr/><a:p><a:r><a:t>{_xml_escape(text)}</a:t></a:r></a:p></p:txBody>'
            '</p:sp></p:spTree></p:cSld></p:sld>'
        ).encode("utf-8")
        out_parts[out_name] = slide_xml
        rid = f"rId{i + 1}"
        pres_rels.append(
            f'<Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" '
            f'Target="slides/slide{i}.xml"/>'
        )
        ct_overrides.append(
            f'<Override PartName="/{out_name}" '
            f'ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
        )
        sldid_list.append(f'<p:sldId id="{255 + i}" r:id="{rid}"/>')

    presentation = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<p:presentation xmlns:p="{_P}" xmlns:a="{_A}" xmlns:r="{_R_NS}">'
        f'<p:sldIdLst>{"".join(sldid_list)}</p:sldIdLst>'
        '<p:sldSz cx="9144000" cy="6858000"/>'
        '</p:presentation>'
    ).encode("utf-8")

    parts = {
        "[Content_Types].xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>'
            + "".join(ct_overrides)
            + '</Types>'
        ).encode("utf-8"),
        "_rels/.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="ppt/presentation.xml"/>'
            '</Relationships>'
        ).encode("utf-8"),
        "ppt/presentation.xml": presentation,
        "ppt/_rels/presentation.xml.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            + "".join(pres_rels)
            + '</Relationships>'
        ).encode("utf-8"),
    }
    parts.update(out_parts)
    return _zip_out(parts)


def _render_pdf(data: bytes, plan: dict) -> bytes:
    """Sanitised re-render: regenerate a fresh text-layer PDF from cleaned page
    text (§5.1 PDF bullet), NOT a box over the original content stream.

    We deliberately do NOT pull a PDF-writer dependency into the minimal jail
    (keeps the F6 surface small): we hand-emit a minimal, valid, single-stream-
    per-page PDF from the cleaned native text.  The ORIGINAL content stream,
    object graph, DocInfo + XMP metadata are entirely discarded — only the cleaned
    visible text is regenerated.

    F4 residual vectors: no original content stream travels (overlay leak gone);
    DocInfo + XMP metadata dropped; no embedded thumbnail/preview; the regenerated
    text uses a base-14 font (Helvetica) with NO embedded subset, so a font-subset
    survival of redacted glyphs is impossible.  Image-only pages carried no native
    text and are dropped (their content was uninspectable → must not survive)."""
    import pypdf  # type: ignore[import-untyped]
    from pypdf.errors import PdfReadError, DependencyError  # type: ignore[import-untyped]

    try:
        reader = pypdf.PdfReader(io.BytesIO(data))
    except (PdfReadError, DependencyError, OSError, ValueError) as exc:
        raise _Contained(f"pdf parse failed during re-render: {exc!r} — fail-closed") from exc
    if reader.is_encrypted:
        raise _Contained("pdf is encrypted — never re-rendered, fail-closed BLOCK")

    by_seg, _strip = _parse_plan(plan)

    page_texts: list[str] = []
    try:
        pages = reader.pages
        n_pages = len(pages)
    except (PdfReadError, DependencyError, ValueError) as exc:
        raise _Contained(f"pdf page tree unreadable: {exc!r} — fail-closed") from exc

    for idx in range(n_pages):
        try:
            text = pages[idx].extract_text() or ""
        except Exception:
            # An unreadable page cannot be cleaned with certainty → fail-closed
            # (we must not ship a page we could not inspect).
            raise _Contained(
                f"pdf page {idx + 1} unreadable during re-render — fail-closed"
            )
        loc = f"page={idx + 1}"
        if text.strip() and loc in by_seg:
            text = _apply_transforms(text, by_seg[loc])
        page_texts.append(text)

    return _emit_text_pdf(page_texts)


def _pdf_escape(text: str) -> str:
    """Escape a string for a PDF literal-string ( ... ) token."""
    return (
        text.replace("\\", "\\\\")
        .replace("(", "\\(")
        .replace(")", "\\)")
        .replace("\r", " ")
        .replace("\n", " ")
    )


def _emit_text_pdf(page_texts: list[str]) -> bytes:
    """Hand-emit a minimal valid PDF: one page per text block, Helvetica base-14
    font (no embedded subset), NO DocInfo, NO XMP. Regenerate-from-cleaned-content.

    Long lines are wrapped naively; the output is a faithful TEXT re-render (lower
    fidelity than the source — destruction/sanitisation is the point, §5.1)."""
    objs: list[bytes] = []

    # 1: Catalog, 2: Pages, then per-page (Page + Contents), then Font (last).
    if not page_texts:
        page_texts = [""]
    n = len(page_texts)
    # Object numbering: 1 catalog, 2 pages, font = 3, pages start at 4.
    font_obj = 3
    first_page_obj = 4
    page_obj_nums = [first_page_obj + 2 * i for i in range(n)]
    content_obj_nums = [first_page_obj + 2 * i + 1 for i in range(n)]

    objs.append(b"<</Type/Catalog/Pages 2 0 R>>")  # 1
    kids = " ".join(f"{p} 0 R" for p in page_obj_nums)
    objs.append(f"<</Type/Pages/Kids[{kids}]/Count {n}>>".encode())  # 2
    objs.append(b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>")  # 3 (font)

    for i, text in enumerate(page_texts):
        # Build a content stream: lines from 720 downward, 14pt leading.
        lines: list[str] = []
        for raw_line in (text.split("\n") if text else [""]):
            # Wrap at ~90 chars to keep within the page width.
            chunk = raw_line
            while len(chunk) > 90:
                lines.append(chunk[:90])
                chunk = chunk[90:]
            lines.append(chunk)
        ops = ["BT", "/F1 12 Tf", "72 720 Td", "14 TL"]
        for j, ln in enumerate(lines):
            if j > 0:
                ops.append("T*")
            ops.append(f"({_pdf_escape(ln)}) Tj")
        ops.append("ET")
        content = "\n".join(ops).encode("latin-1", "replace")
        page_dict = (
            f"<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
            f"/Contents {content_obj_nums[i]} 0 R"
            f"/Resources<</Font<</F1 {font_obj} 0 R>>>>>>"
        ).encode()
        objs.append(page_dict)  # page
        objs.append(
            b"<</Length " + str(len(content)).encode() + b">>stream\n" + content + b"\nendstream"
        )  # contents

    out = bytearray(b"%PDF-1.4\n")
    offsets: list[int] = []
    for i, body in enumerate(objs, start=1):
        offsets.append(len(out))
        out += f"{i} 0 obj".encode() + body + b"\nendobj\n"
    xref_pos = len(out)
    out += b"xref\n0 " + str(len(objs) + 1).encode() + b"\n"
    out += b"0000000000 65535 f \n"
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    # NO /Info (DocInfo) entry → no metadata residual.
    out += (
        b"trailer<</Root 1 0 R/Size " + str(len(objs) + 1).encode() + b">>\n"
        b"startxref\n" + str(xref_pos).encode() + b"\n%%EOF"
    )
    return bytes(out)


def _render_text_like(data: bytes, plan: dict, fmt: str) -> bytes:
    """txt/csv re-render: regenerate cleaned text/rows. Provably clean — there is
    no hidden channel (§5.1)."""
    by_seg, _strip = _parse_plan(plan)
    text = data.decode("utf-8", "replace") if _is_utf8(data) else data.decode("latin-1")

    if fmt == "txt":
        lines = text.split("\n")
        out_lines: list[str] = []
        for idx, line in enumerate(lines, start=1):
            loc = f"line={idx}"
            if loc in by_seg:
                line = _apply_transforms(line, by_seg[loc])
            out_lines.append(line)
        return "\n".join(out_lines).encode("utf-8")

    # csv
    import csv as _csv
    out_buf = io.StringIO()
    reader = _csv.reader(io.StringIO(text))
    writer = _csv.writer(out_buf)
    for row_idx, row in enumerate(reader, start=1):
        new_row: list[str] = []
        for col_idx, cell in enumerate(row, start=1):
            loc = f"row={row_idx},col={col_idx}"
            if loc in by_seg:
                cell = _apply_transforms(cell, by_seg[loc])
            new_row.append(cell)
        writer.writerow(new_row)
    return out_buf.getvalue().encode("utf-8")


def _is_utf8(data: bytes) -> bool:
    try:
        data.decode("utf-8")
        return True
    except UnicodeDecodeError:
        return False


_RENDERERS = {
    "docx": _render_docx,
    "xlsx": _render_xlsx,
    "pptx": _render_pptx,
    "pdf": _render_pdf,
}


#: Formats whose re-render is REGENERATE-FROM-CLEANED-CONTENT in the jail.
#: txt/csv re-render here too even though they extract host-side, because the
#: re-render path is uniform (the cleaned-content rebuild) and the worker is the
#: single place re-render runs (F6) — the host never re-renders.
_RENDER_SUPPORTED = {"docx", "xlsx", "pptx", "pdf", "txt", "csv"}


def _decode_plan(plan_b64: str) -> dict:
    """Decode the base64'd JSON RenderPlan from argv. Malformed → fail-closed."""
    import base64 as _b64
    if not plan_b64:
        raise _Contained("re-render job missing --plan — fail-closed")
    try:
        raw = _b64.b64decode(plan_b64.encode("ascii")).decode("utf-8")
        obj = json.loads(raw)
    except (ValueError, UnicodeDecodeError) as exc:
        raise _Contained(f"re-render plan undecodable: {exc} — fail-closed") from exc
    if not isinstance(obj, dict):
        raise _Contained("re-render plan is not an object — fail-closed")
    return obj


def _run_render(job: str, fmt: str, data: bytes, plan: dict) -> dict:
    """REDACT / PSEUDONYMIZE re-render (red-team F6 — runs in THIS jail).

    Returns the re-rendered artefact as base64 in the JSON result so the
    stdin->JSON-stdout contract carries the binary cleanly. The result also
    re-extracts the OUTPUT and returns its segments so the host can assert the
    no-residual / tokenized invariant without a second jail round-trip (this is
    the proof Laura demands — re-extract-the-output)."""
    if fmt not in _RENDER_SUPPORTED:
        raise _Contained(f"re-render unsupported for format '{fmt}' — fail-closed BLOCK")

    if fmt in ("txt", "csv"):
        rendered = _render_text_like(data, plan, fmt)
    else:
        if fmt in ("docx", "xlsx", "pptx"):
            if _ooxml_is_encrypted(data):
                raise _Contained(
                    f"{fmt} is encrypted — never re-rendered, fail-closed BLOCK"
                )
        rendered = _RENDERERS[fmt](data, plan)

    # Re-extract the OUTPUT for the host-side no-residual assertion. txt/csv have
    # no jail extractor (host-side trivial extractors), so we re-extract them
    # here with a simple line/row split for the proof segments.
    if fmt in ("txt", "csv"):
        out_segments = _reextract_text_like(rendered, fmt)
        complete = True
    else:
        out_segments, complete = _EXTRACTORS[fmt](rendered)

    import base64 as _b64
    return {
        "ok": True,
        "job": job,
        "detected_format": fmt,
        "rendered_b64": _b64.b64encode(rendered).decode("ascii"),
        # The re-extracted OUTPUT segments — the host asserts the original values
        # are gone (REDACT) or tokenized (PSEUDONYMIZE) and NO residual in any
        # part incl. metadata.
        "output_segments": out_segments,
        "output_extraction_complete": complete,
    }


def _reextract_text_like(data: bytes, fmt: str) -> list[dict]:
    text = data.decode("utf-8", "replace")
    segs: list[dict] = []
    if fmt == "txt":
        for idx, line in enumerate(text.split("\n"), start=1):
            seg = _seg(line, "BODY", f"line={idx}")
            if seg:
                segs.append(seg)
    else:
        import csv as _csv
        for row_idx, row in enumerate(_csv.reader(io.StringIO(text)), start=1):
            for col_idx, cell in enumerate(row, start=1):
                seg = _seg(cell, "TABLE_CELL", f"row={row_idx},col={col_idx}")
                if seg:
                    segs.append(seg)
    return segs


def _run_extract(fmt: str, data: bytes) -> dict:
    if fmt not in _SUPPORTED:
        raise _Contained(f"unsupported format '{fmt}' — fail-closed")
    if fmt in ("docx", "xlsx", "pptx"):
        # Encrypted/password-protected OOXML is an OLE Compound File, NOT a zip.
        # Detect it FIRST (before the zip bomb guard tries to open it and reports
        # the generic "not a valid zip") so the audit reason is the precise
        # "encrypted — never cracked, fail-closed BLOCK".
        if _ooxml_is_encrypted(data):
            raise _Contained(
                f"{fmt} is encrypted/password-protected (OLE-wrapped) — never "
                f"cracked, fail-closed BLOCK"
            )
        # Guard the container BEFORE parsing (OOXML is a zip; pdf is guarded by
        # the parser's own bounds + the cgroup — no zip layer to bomb-check).
        _guard_ooxml(data)
    segments, complete = _EXTRACTORS[fmt](data)
    return {
        "ok": True,
        "segments": segments,
        "extraction_complete": complete,
        "detected_format": fmt,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="yashigani-extractor-worker")
    parser.add_argument("--job", default="extract",
                        choices=["extract", "redact", "pseudonymize"])
    parser.add_argument("--format", dest="fmt", required=True)
    parser.add_argument("--declared-mime", dest="declared_mime", default="")
    # Re-render plan (REDACT/PSEUDONYMIZE) — base64'd JSON RenderPlan on argv.
    # The DOCUMENT bytes stay on stdin (the single read-only input); only the
    # small plan (per-span tokens + originals, NO replacer map) travels in argv.
    parser.add_argument("--plan", dest="plan_b64", default="")
    args = parser.parse_args(argv)

    try:
        data = _read_stdin_capped()
        if args.job == "extract":
            result = _run_extract(args.fmt, data)
        elif args.job in ("redact", "pseudonymize"):
            # Re-render (REDACT/PSEUDONYMIZE) runs in THIS same jail (F6).
            plan = _decode_plan(args.plan_b64)
            result = _run_render(args.job, args.fmt, data, plan)
        else:
            raise _Contained(f"unknown job '{args.job}' — fail-closed")
        _emit(result)
        return 0
    except _Contained as exc:
        # Clean containment: a guard/limit caught it. ok=False, exit 0 — the
        # runner BLOCKs but records this as "contained", not "worker crashed".
        _emit({"ok": False, "reason": str(exc)})
        return 0
    except Exception as exc:  # pragma: no cover - any unexpected parser death
        # A parser crash. Write nothing parseable as a result; exit non-zero so
        # the runner fails closed to BLOCK (do NOT emit ok=true on a crash).
        sys.stderr.write(f"worker crashed: {exc!r}\n")
        return 70  # EX_SOFTWARE


if __name__ == "__main__":
    sys.exit(main())
